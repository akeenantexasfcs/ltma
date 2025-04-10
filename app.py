#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import io
import json
import os
import re
import tempfile
import time

# Third-party libraries
import boto3
import botocore
import numpy as np
import pandas as pd
import streamlit as st
from Levenshtein import distance as levenshtein_distance
from openpyxl import load_workbook

# Module configuration
import logging
logging.basicConfig(level=logging.INFO)

# Set up logging
logging.basicConfig(level=logging.INFO)

# General Utility Functions
def process_file(file):
    logging.info(f"Processing file: {file.name}")
    try:
        df = pd.read_excel(file, sheet_name=None)
        first_sheet_name = list(df.keys())[0]
        df = df[first_sheet_name]
        return df
    except Exception as e:
        st.error(f"Error processing file {file.name}: {e}")
        return None

def create_combined_df(dfs):
    logging.info("Combining data frames...")
    combined_df = pd.DataFrame()
    for i, df in enumerate(dfs):
        final_mnemonic_col = 'Final Mnemonic Selection'
        if final_mnemonic_col not in df.columns:
            st.error(f"Column '{final_mnemonic_col}' not found in dataframe {i+1}")
            continue

        date_cols = [col for col in df.columns if col not in ['Label', 'Account', final_mnemonic_col, 'Mnemonic', 'Manual Selection']]
        if not date_cols:
            st.error(f"No date columns found in dataframe {i+1}")
            continue

        df_grouped = df.groupby([final_mnemonic_col, 'Label']).sum(numeric_only=True).reset_index()
        df_melted = df_grouped.melt(id_vars=[final_mnemonic_col, 'Label'], value_vars=date_cols, var_name='Date', value_name='Value')
        df_pivot = df_melted.pivot(index=['Label', final_mnemonic_col], columns='Date', values='Value')

        if combined_df.empty:
            combined_df = df_pivot
        else:
            combined_df = combined_df.join(df_pivot, how='outer')
    return combined_df.reset_index()

def aggregate_data(df):
    logging.info("Aggregating data...")
    if 'Label' not in df.columns or 'Account' not in df.columns:
        st.error("'Label' and/or 'Account' columns not found in the data.")
        return df

    pivot_table = df.pivot_table(index=['Label', 'Account'], 
                                 values=[col for col in df.columns if col not in ['Label', 'Account', 'Mnemonic', 'Manual Selection']], 
                                 aggfunc='sum').reset_index()
    return pivot_table

def clean_numeric_value(value):
    try:
        value_str = str(value).strip()
        
        # Remove any number of spaces between $ and (
        value_str = re.sub(r'\$\s*\(', '$(', value_str)
        
        # Handle negative values in parentheses with or without dollar sign
        if (value_str.startswith('$(') and value_str.endswith(')')) or (value_str.startswith('(') and value_str.endswith(')')):
            value_str = '-' + value_str.lstrip('$(').rstrip(')')
        
        # Remove dollar signs and commas
        cleaned_value = re.sub(r'[$,]', '', value_str)
        
        # Convert text to number
        try:
            from word2number import w2n
            cleaned_value = w2n.word_to_num(cleaned_value)
        except (ValueError, ImportError):
            pass
        
        return float(cleaned_value)
    except (ValueError, TypeError) as e:
        logging.error(f"Error converting value: {value} with error: {e}")
        return 0

def save_lookup_table(df, file_path):
    logging.info("Writing data to Excel...")
    df.to_excel(file_path, index=False)

def apply_unit_conversion(df, columns, factor):
    logging.info("Applying unit conversion...")
    for selected_column in columns:
        if selected_column in df.columns:
            df[selected_column] = df[selected_column].apply(
                lambda x: x * factor if isinstance(x, (int, float)) else x)
    return df

def check_all_zeroes(df):
    logging.info("Checking for all zeroes in data frame...")
    zeroes = (df.iloc[:, 2:] == 0).all(axis=1)
    return zeroes

def get_best_match(account_value, label_value, lookup_df, threshold=0.4):
    """
    Improved fuzzy matching function to find the best match in the lookup table
    
    Parameters:
    -----------
    account_value : str
        The account name to match
    label_value : str
        The label/category of the account
    lookup_df : pandas.DataFrame
        The lookup dictionary dataframe
    threshold : float
        The similarity threshold (0-1) below which a match is considered valid
        
    Returns:
    --------
    tuple
        (best_match_row, score) - The best matching row and its similarity score
    """
    # First filter by label if available
    filtered_df = lookup_df
    if label_value and not pd.isna(label_value) and 'Label' in lookup_df.columns:
        label_matches = lookup_df[lookup_df['Label'].str.lower() == str(label_value).lower()]
        if not label_matches.empty:
            filtered_df = label_matches
    
    # If no account value, return no match
    if pd.isna(account_value) or not account_value:
        return None, float('inf')
        
    # Calculate similarity scores
    best_score = float('inf')
    best_match = None
    
    for _, row in filtered_df.iterrows():
        lookup_account = str(row['Account']).lower()
        account_str = str(account_value).lower()
        
        # Calculate normalized Levenshtein distance
        score = levenshtein_distance(account_str, lookup_account) / max(len(account_str), len(lookup_account))
        
        # Calculate token-based similarity (partial matches)
        account_tokens = set(account_str.split())
        lookup_tokens = set(lookup_account.split())
        common_tokens = account_tokens.intersection(lookup_tokens)
        
        # Token similarity score (0-1, lower is better)
        if account_tokens and lookup_tokens:
            token_score = 1 - len(common_tokens) / max(len(account_tokens), len(lookup_tokens))
        else:
            token_score = 1
        
        # Combined score (weighted average)
        combined_score = 0.7 * score + 0.3 * token_score
        
        if combined_score < best_score:
            best_score = combined_score
            best_match = row
    
    # Return match if score is below threshold
    if best_score < threshold:
        return best_match, best_score
    else:
        return None, best_score
    
def balance_sheet_BS():
    # Initial setup and data loading
    balance_sheet_data_dictionary_file = 'balance_sheet_data_dictionary.xlsx'
    
    # Define the initial lookup data for Balance Sheet
    initial_balance_sheet_lookup_data = {
        "Label": ["Current Assets", "Non Current Assets", "Non Current Assets", "Non Current Assets", "Non Current Assets", "Non Current Assets", "Non Current Assets"],
        "Account": ["Gross Property, Plant & Equipment", "Accumulated Depreciation", "Net Property, Plant & Equipment", "Long-term Investments", "Goodwill", "Other Intangibles", "Right-of-Use Asset-Net"],
        "Mnemonic": ["Gross Property, Plant & Equipment", "Accumulated Depreciation", "Net Property, Plant & Equipment", "Long-term Investments", "Goodwill", "Other Intangibles", "Right-of-Use Asset-Net"],
        "CIQ": ["IQ_GPPE", "IQ_AD", "IQ_NPPE", "IQ_LT_INVEST", "IQ_GW", "IQ_OTHER_INTAN", "IQ_RUA_NET"]
    }
    
    # Load or initialize the lookup table
    @st.cache_data(ttl=600)
    def load_balance_sheet_data():
        logging.info("Loading or initializing balance sheet data...")
        try:
            return pd.read_excel(balance_sheet_data_dictionary_file)
        except FileNotFoundError:
            return pd.DataFrame(initial_balance_sheet_lookup_data)

    if 'balance_sheet_data' not in st.session_state:
        st.session_state.balance_sheet_data = load_balance_sheet_data()

    def save_and_update_balance_sheet_data(df):
        logging.info("Saving and updating balance sheet data...")
        st.session_state.balance_sheet_data = df
        save_lookup_table(df, balance_sheet_data_dictionary_file)
    
    def sort_by_label_and_account(df):
        logging.info("Sorting by label and account...")
        sort_order = {
            "Current Assets": 0,
            "Non Current Assets": 1,
            "Current Liabilities": 2,
            "Non Current Liabilities": 3,
            "Equity": 4,
            "Total Equity and Liabilities": 5
        }
        df['Label_Order'] = df['Label'].map(sort_order)
        df['Total_Order'] = df['Account'].str.contains('Total', case=False).astype(int)
        df = df.sort_values(by=['Label_Order', 'Label', 'Total_Order', 'Account']).drop(columns=['Label_Order', 'Total_Order'])
        return df

    def sort_by_label_and_final_mnemonic(df):
        logging.info("Sorting by label and final mnemonic selection...")
        sort_order = {
            "Current Assets": 0,
            "Non Current Assets": 1,
            "Current Liabilities": 2,
            "Non Current Liabilities": 3,
            "Equity": 4,
            "Total Equity and Liabilities": 5
        }
        df['Label_Order'] = df['Label'].map(sort_order)
        df['Total_Order'] = df['Final Mnemonic Selection'].str.contains('Total', case=False).astype(int)
        df = df.sort_values(by=['Label_Order', 'Total_Order', 'Final Mnemonic Selection']).drop(columns=['Label_Order', 'Total_Order'])
        return df
    
    st.title("BALANCE SHEET LTMA")

    tab1, tab2, tab3, tab4 = st.tabs(["Table Extractor", "Aggregate My Data", "Mappings and Data Consolidation", "Balance Sheet Data Dictionary"])

    with tab1:
        uploaded_file = st.file_uploader("Choose a JSON file", type="json", key='json_uploader')

        # Persist uploaded file across sessions
        if uploaded_file is not None:
            st.session_state.uploaded_file = uploaded_file

        if 'uploaded_file' in st.session_state:
            uploaded_file = st.session_state.uploaded_file
            
            try:
                data = json.load(uploaded_file)
            except json.JSONDecodeError:
                st.error("The uploaded file is not a valid JSON. Please upload a valid JSON file.")
                return
            
            st.warning("PLEASE NOTE: In the Setting Bounds Preview Window, you will see only your respective labels. In the Updated Columns Preview Window, you will see only your renamed column headers. The labels from the Setting Bounds section will not appear in the Updated Columns Preview.")
            st.warning("PLEASE ALSO NOTE: An Account column must also be designated when you are in the Rename Columns section.")

            tables = []
            for block in data['Blocks']:
                if block['BlockType'] == 'TABLE':
                    table = {}
                    if 'Relationships' in block:
                        for relationship in block['Relationships']:
                            if relationship['Type'] == 'CHILD':
                                for cell_id in relationship['Ids']:
                                    cell_block = next((b for b in data['Blocks'] if b['Id'] == cell_id), None)
                                    if cell_block:
                                        row_index = cell_block.get('RowIndex', 0)
                                        col_index = cell_block.get('ColumnIndex', 0)
                                        if row_index not in table:
                                            table[row_index] = {}
                                        cell_text = ''
                                        if 'Relationships' in cell_block:
                                            for rel in cell_block['Relationships']:
                                                if rel['Type'] == 'CHILD':
                                                    for word_id in rel['Ids']:
                                                        word_block = next((w for w in data['Blocks'] if w['Id'] == word_id), None)
                                                        if word_block and word_block['BlockType'] == 'WORD':
                                                            cell_text += ' ' + word_block.get('Text', '')
                                        table[row_index][col_index] = cell_text.strip()
                    table_df = pd.DataFrame.from_dict(table, orient='index').sort_index()
                    table_df = table_df.sort_index(axis=1)
                    tables.append(table_df)
            all_tables = pd.concat(tables, axis=0, ignore_index=True)
            if len(all_tables.columns) == 0:
                st.error("No columns found in the uploaded JSON file.")
                return

            column_a = all_tables.columns[0]
            all_tables.insert(0, 'Label', '')

            st.subheader("Data Preview")
            st.dataframe(all_tables)

            def get_unique_options(series):
                """
                Generate a list of unique options for selection, adding index labels for blanks.
                """
                counts = series.value_counts()
                unique_options = []
                occurrence_counts = {}
                for idx, item in series.items():
                    if pd.isna(item) or item == '':
                        unique_options.append(f"[Blank {idx}]")
                    else:
                        if counts[item] > 1:
                            if item not in occurrence_counts:
                                occurrence_counts[item] = 1
                            else:
                                occurrence_counts[item] += 1
                            unique_options.append(f"{item} {occurrence_counts[item]}")
                        else:
                            unique_options.append(item)
                return unique_options

            labels = ["Current Assets", "Non Current Assets", "Current Liabilities",
                      "Non Current Liabilities", "Equity", "Total Equity and Liabilities"]
            selections = []

            for label in labels:
                st.subheader(f"Setting bounds for {label}")
                options = [''] + get_unique_options(all_tables[column_a].fillna(''))
                start_label = st.selectbox(f"Start Label for {label}", options, key=f"start_{label}")
                end_label = st.selectbox(f"End Label for {label}", options, key=f"end_{label}")
                selections.append((label, start_label, end_label))

            new_column_names = {col: col for col in all_tables.columns}

            def update_labels(df):
                df['Label'] = ''
                account_column = new_column_names.get(column_a, column_a)

                # Insert index numbers in blank cells for easier identification
                df[account_column] = df[account_column].apply(lambda x: str(x) if pd.notna(x) else '')

                for index in df.index:
                    if df.at[index, account_column] == '':
                        df.at[index, account_column] = f"[Blank {index}]"

                # Use selections to set label bounds
                for label, start_label, end_label in selections:
                    if start_label or end_label:
                        try:
                            # If start_label is a blank (indicated by index), use index directly
                            if start_label.startswith("[Blank"):
                                start_index = int(re.search(r"\d+", start_label).group())
                            else:
                                start_label_parts = start_label.split()
                                start_label_base = " ".join(start_label_parts[:-1]) if start_label_parts[-1].isdigit() else start_label
                                start_instance = int(start_label_parts[-1]) if start_label_parts[-1].isdigit() else 1
                                start_indices = df[df[account_column].str.contains(start_label_base, regex=False, na=False)].index
                                start_index = start_indices[start_instance - 1] if len(start_indices) >= start_instance else None
                            
                            # If end_label is a blank (indicated by index), use index directly
                            if end_label.startswith("[Blank"):
                                end_index = int(re.search(r"\d+", end_label).group())
                            else:
                                end_label_parts = end_label.split()
                                end_label_base = " ".join(end_label_parts[:-1]) if end_label_parts[-1].isdigit() else end_label
                                end_instance = int(end_label_parts[-1]) if end_label_parts[-1].isdigit() else 1
                                end_indices = df[df[account_column].str.contains(end_label_base, regex=False, na=False)].index
                                end_index = end_indices[end_instance - 1] if len(end_indices) >= end_instance else None
                            
                            # Ensure start and end indices are set, or skip if missing
                            if start_index is not None and end_index is not None:
                                df.loc[start_index:end_index, 'Label'] = label
                            else:
                                st.error(f"Invalid label bounds for {label}. Ensure both start and end labels are properly defined.")

                        except KeyError as e:
                            st.error(f"Error accessing column '{account_column}': {e}. Skipping...")
                    else:
                        st.info(f"No start or end label provided for {label}. Skipping...")

                return df

            if st.button("Preview Setting Bounds ONLY", key="preview_setting_bounds"):
                preview_table = update_labels(all_tables.copy())
                st.subheader("Preview of Setting Bounds")
                st.dataframe(preview_table)

            st.subheader("Rename Columns")
            new_column_names = {}
            fiscal_year_options = [f"FY{year}" for year in range(2018, 2027)]
            ytd_options = [f"YTD{quarter}{year}" for year in range(2018, 2027) for quarter in range(1, 4)]
            dropdown_options = [''] + ['Account'] + fiscal_year_options + ytd_options

            for col in all_tables.columns:
                new_name_text = st.text_input(f"Rename '{col}' to:", value=col, key=f"rename_{col}_text")
                new_name_dropdown = st.selectbox(f"Or select predefined name for '{col}':", dropdown_options, key=f"rename_{col}_dropdown", index=0)
                new_column_names[col] = new_name_dropdown if new_name_dropdown else new_name_text

            all_tables.rename(columns=new_column_names, inplace=True)
            st.write("Updated Columns:", all_tables.columns.tolist())
            st.dataframe(all_tables)

            st.subheader("Edit Data Frame")
            editable_df = st.data_editor(all_tables)
            st.write("Editable Data Frame:", editable_df)

            st.subheader("Select columns to keep before export")
            columns_to_keep = []
            for col in editable_df.columns:
                if st.checkbox(f"Keep column '{col}'", value=True, key=f"keep_{col}"):
                    columns_to_keep.append(col)

            st.subheader("Select numerical columns")
            numerical_columns = []
            for col in editable_df.columns:
                if st.checkbox(f"Numerical column '{col}'", value=False, key=f"num_{col}"):
                    numerical_columns.append(col)

            if 'Label' not in columns_to_keep:
                columns_to_keep.insert(0, 'Label')

            if 'Account' not in columns_to_keep:
                columns_to_keep.insert(1, 'Account')

            st.subheader("Label Units")
            selected_columns = st.multiselect("Select columns for conversion", options=numerical_columns, key="columns_selection")
            selected_value = st.radio("Select conversion value", ["Actuals", "Thousands", "Millions", "Billions"], index=0, key="conversion_value")

            conversion_factors = {
                "Actuals": 1,
                "Thousands": 1000,
                "Millions": 1000000,
                "Billions": 1000000000
            }

            if st.button("Apply Selected Labels and Generate Excel", key="apply_selected_labels_generate_excel_tab1"):
                updated_table = update_labels(editable_df.copy())
                updated_table = updated_table[[col for col in columns_to_keep if col in updated_table.columns]]

                updated_table = updated_table[updated_table['Label'].str.strip() != '']
                updated_table = updated_table[updated_table['Account'].str.strip() != '']

                for col in numerical_columns:
                    if col in updated_table.columns:
                        updated_table[col] = updated_table[col].apply(clean_numeric_value)

                if selected_value != "Actuals":
                    updated_table = apply_unit_conversion(updated_table, selected_columns, conversion_factors[selected_value])

                updated_table.replace('-', 0, inplace=True)

                excel_file = io.BytesIO()
                updated_table.to_excel(excel_file, index=False)
                excel_file.seek(0)
                st.download_button("Download Excel", excel_file, "Table_Extractor_Balance_Sheet.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            st.subheader("Check for Duplicate Accounts")
            if 'Account' not in editable_df.columns:
                st.warning("The 'Account' column is missing. Please ensure your data includes an 'Account' column.")
            else:
                duplicated_accounts = editable_df[editable_df.duplicated(['Account'], keep=False)]
                if not duplicated_accounts.empty:
                    st.warning("Duplicates identified:")
                    st.dataframe(duplicated_accounts)
                else:
                    st.success("No duplicates identified")

    with tab2:
        st.subheader("Aggregate My Data")

        uploaded_files = st.file_uploader("Upload your Excel files from Tab 1", type=['xlsx'], accept_multiple_files=True, key='xlsx_uploader_tab2')

        dfs = []
        if uploaded_files:
            dfs = [process_file(file) for file in uploaded_files if process_file(file) is not None]

        if dfs:
            combined_df = pd.concat(dfs, ignore_index=True)
            st.dataframe(combined_df)

            aggregated_table = aggregate_data(combined_df)
            aggregated_table = sort_by_label_and_account(aggregated_table)

            st.subheader("Aggregated Data")
            st.dataframe(aggregated_table)

            st.subheader("Preview Data and Edit Rows")
            zero_rows = check_all_zeroes(aggregated_table)
            zero_rows_indices = aggregated_table.index[zero_rows].tolist()
            st.write("Rows where all values (past the first 2 columns) are zero:", aggregated_table.loc[zero_rows_indices])

            edited_data = st.data_editor(aggregated_table, num_rows="dynamic")

            st.write("Highlighted rows with all zero values for potential removal:")
            for index in zero_rows_indices:
                st.write(f"Row {index}: {aggregated_table.loc[index].to_dict()}")

            rows_removed = False
            if st.button("Remove Highlighted Rows", key="remove_highlighted_rows"):
                aggregated_table = aggregated_table.drop(zero_rows_indices).reset_index(drop=True)
                rows_removed = True
                st.success("Highlighted rows removed successfully")
                st.dataframe(aggregated_table)

            st.subheader("Download Aggregated Data")
            download_label = "Download Updated Aggregated Excel" if rows_removed else "Download Aggregated Excel"
            excel_file = io.BytesIO()
            with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
                aggregated_table.to_excel(writer, sheet_name='Aggregated Data', index=False)
            excel_file.seek(0)
            st.download_button(download_label, excel_file, "Aggregate_My_Data_Balance_Sheet.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("Please upload valid Excel files for aggregation.")

    with tab3:
        st.subheader("Mappings and Data Consolidation")

        uploaded_excel = st.file_uploader("Upload your Excel file for Mnemonic Mapping", type=['xlsx'], key='excel_uploader_tab3_bs')

        currency_options = ["U.S. Dollar", "Euro", "British Pound Sterling", "Japanese Yen"]
        magnitude_options = ["Actuals", "Thousands", "Millions", "Billions", "Trillions"]

        selected_currency = st.selectbox("Select Currency", currency_options, key='currency_selection_tab3_bs')
        selected_magnitude = st.selectbox("Select Magnitude", magnitude_options, key='magnitude_selection_tab3_bs')
        company_name_bs = st.text_input("Enter Company Name", key='company_name_input_bs')

        if uploaded_excel is not None:
            df = pd.read_excel(uploaded_excel)

            statement_dates = {}
            for col in df.columns[2:]:
                statement_date = st.text_input(f"Enter statement date for {col}", key=f"statement_date_{col}_bs")
                statement_dates[col] = statement_date

            st.write("Columns in the uploaded file:", df.columns.tolist())

            if 'Account' not in df.columns:
                st.error("The uploaded file does not contain an 'Account' column.")
            else:
                # Apply the new fuzzy matching function
                df['Mnemonic'] = ''
                df['Manual Selection'] = ''
                for idx, row in df.iterrows():
                    account_value = row['Account']
                    label_value = row.get('Label', '')
                    if pd.notna(account_value):
                        best_match, score = get_best_match(account_value, label_value, st.session_state.balance_sheet_data)
                        if best_match is not None:
                            df.at[idx, 'Mnemonic'] = best_match['Mnemonic']
                            # Show match score for debugging
                            st.write(f"Match for '{account_value}': {best_match['Mnemonic']} (Score: {score:.2f})")
                        else:
                            df.at[idx, 'Mnemonic'] = 'Manual Mapping Required'
                            st.markdown(f"**Manual Mapping Required for:** {account_value} [{label_value} - Index {idx}]")

                for idx, row in df.iterrows():
                    account_value = row['Account']
                    label_value = row.get('Label', '')
                    
                    # If manual mapping required, show dropdown with all possible mnemonics
                    if row['Mnemonic'] == 'Manual Mapping Required':
                        # Get mnemonics for the specific label if available
                        if pd.notna(label_value) and label_value != '':
                            label_mnemonics = st.session_state.balance_sheet_data[
                                st.session_state.balance_sheet_data['Label'].str.strip().str.lower() == label_value.strip().lower()
                            ]['Mnemonic'].unique()
                            
                            # If no mnemonics for this label, show all mnemonics
                            if len(label_mnemonics) == 0:
                                label_mnemonics = st.session_state.balance_sheet_data['Mnemonic'].unique()
                        else:
                            label_mnemonics = st.session_state.balance_sheet_data['Mnemonic'].unique()
                        
                        manual_selection_options = list(label_mnemonics)
                        manual_selection = st.selectbox(
                            f"Select mapping for '{account_value}'",
                            options=[''] + manual_selection_options + ['REMOVE ROW', 'MANUAL OVERRIDE'],
                            key=f"select_{idx}_tab3_bs"
                        )
                        if manual_selection:
                            df.at[idx, 'Manual Selection'] = manual_selection.strip()
                    else:
                        # If automatic mapping worked, still allow override
                        manual_override = st.checkbox(f"Override mapping for '{account_value}'", key=f"override_{idx}_tab3_bs")
                        if manual_override:
                            label_mnemonics = st.session_state.balance_sheet_data['Mnemonic'].unique()
                            manual_selection_options = list(label_mnemonics)
                            manual_selection = st.selectbox(
                                f"Select mapping for '{account_value}'",
                                options=[''] + manual_selection_options + ['REMOVE ROW'],
                                key=f"select_override_{idx}_tab3_bs"
                            )
                            if manual_selection:
                                df.at[idx, 'Manual Selection'] = manual_selection.strip()

                st.dataframe(df[['Label', 'Account', 'Mnemonic', 'Manual Selection']])

                if st.button("Generate Excel with Lookup Results", key="generate_excel_lookup_results_tab3_bs"):
                    df['Final Mnemonic Selection'] = df.apply(
                        lambda row: row['Manual Selection'] if row['Manual Selection'] != '' else row['Mnemonic'], 
                        axis=1
                    )
                    final_output_df = df[df['Manual Selection'] != 'REMOVE ROW'].copy()

                    combined_df = create_combined_df([final_output_df])
                    combined_df = sort_by_label_and_final_mnemonic(combined_df)

                    def lookup_ciq(mnemonic):
                        if mnemonic == 'Manual Mapping Required':
                            return 'CIQ ID Required'
                        ciq_value = st.session_state.balance_sheet_data.loc[st.session_state.balance_sheet_data['Mnemonic'] == mnemonic, 'CIQ']
                        if ciq_value.empty:
                            return 'CIQ ID Required'
                        return ciq_value.values[0]

                    combined_df['CIQ'] = combined_df['Final Mnemonic Selection'].apply(lookup_ciq)

                    columns_order = ['Label', 'Final Mnemonic Selection', 'CIQ'] + [col for col in combined_df.columns if col not in ['Label', 'Final Mnemonic Selection', 'CIQ']]
                    combined_df = combined_df[columns_order]

                    as_presented_df = final_output_df.drop(columns=['CIQ', 'Mnemonic', 'Manual Selection'], errors='ignore')
                    as_presented_columns_order = ['Label', 'Account', 'Final Mnemonic Selection'] + [col for col in as_presented_df.columns if col not in ['Label', 'Account', 'Final Mnemonic Selection']]
                    as_presented_df = as_presented_df[as_presented_columns_order]

                    excel_file = io.BytesIO()
                    with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
                        combined_df.to_excel(writer, sheet_name='Standardized - Balance Sheet', index=False)
                        as_presented_df.to_excel(writer, sheet_name='As Presented - Balance Sheet', index=False)
                        cover_df = pd.DataFrame({
                            'Selection': ['Currency', 'Magnitude', 'Company Name'] + list(statement_dates.keys()),
                            'Value': [selected_currency, selected_magnitude, company_name_bs] + list(statement_dates.values())
                        })
                        cover_df.to_excel(writer, sheet_name='Cover', index=False)
                    excel_file.seek(0)
                    st.download_button("Download Excel", excel_file, "Mappings_and_Data_Consolidation_Balance_Sheet.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                if st.button("Update Data Dictionary with Manual Mappings", key="update_data_dictionary_tab3_bs"):
                    df['Final Mnemonic Selection'] = df.apply(
                        lambda row: row['Manual Selection'] if row['Manual Selection'] != '' else row['Mnemonic'], 
                        axis=1
                    )
                    new_entries = []
                    for idx, row in df.iterrows():
                        manual_selection = row['Manual Selection']
                        final_mnemonic = row['Final Mnemonic Selection']
                        ciq_value = st.session_state.balance_sheet_data.loc[st.session_state.balance_sheet_data['Mnemonic'] == final_mnemonic, 'CIQ'].values[0] if not st.session_state.balance_sheet_data.loc[st.session_state.balance_sheet_data['Mnemonic'] == final_mnemonic, 'CIQ'].empty else 'CIQ ID Required'

                        if manual_selection == 'REMOVE ROW':
                            st.session_state.balance_sheet_data = st.session_state.balance_sheet_data[st.session_state.balance_sheet_data['Account'] != row['Account']]
                        elif manual_selection != '':
                            if row['Account'] not in st.session_state.balance_sheet_data['Account'].values:
                                new_entries.append({'Account': row['Account'], 'Mnemonic': final_mnemonic, 'CIQ': ciq_value, 'Label': row['Label']})
                            else:
                                st.session_state.balance_sheet_data.loc[st.session_state.balance_sheet_data['Account'] == row['Account'], 'Mnemonic'] = final_mnemonic
                                st.session_state.balance_sheet_data.loc[st.session_state.balance_sheet_data['Account'] == row['Account'], 'Label'] = row['Label']
                                st.session_state.balance_sheet_data.loc[st.session_state.balance_sheet_data['Account'] == row['Account'], 'CIQ'] = ciq_value
                    if new_entries:
                        st.session_state.balance_sheet_data = pd.concat([st.session_state.balance_sheet_data, pd.DataFrame(new_entries)], ignore_index=True)
                    st.session_state.balance_sheet_data.reset_index(drop=True, inplace=True)
                    save_and_update_balance_sheet_data(st.session_state.balance_sheet_data)
                    st.success("Data Dictionary Updated Successfully")
                    st.rerun()

    with tab4:
        st.subheader("Balance Sheet Data Dictionary")

        uploaded_dict_file = st.file_uploader("Upload a new Data Dictionary Excel file", type=['xlsx'], key='dict_uploader_tab4_bs')
        if uploaded_dict_file is not None:
            new_lookup_df = pd.read_excel(uploaded_dict_file)
            st.session_state.balance_sheet_data = new_lookup_df
            save_lookup_table(st.session_state.balance_sheet_data, balance_sheet_data_dictionary_file)
            st.success("Data Dictionary uploaded and updated successfully!")

        st.dataframe(st.session_state.balance_sheet_data)

        remove_indices = st.multiselect("Select rows to remove", st.session_state.balance_sheet_data.index, key='remove_indices_tab4_bs')
        rows_removed = False
        if st.button("Remove Selected Rows", key="remove_selected_rows_tab4_bs"):
            st.session_state.balance_sheet_data = st.session_state.balance_sheet_data.drop(remove_indices).reset_index(drop=True)
            save_lookup_table(st.session_state.balance_sheet_data, balance_sheet_data_dictionary_file)
            rows_removed = True
            st.success("Selected rows removed successfully!")
            st.dataframe(st.session_state.balance_sheet_data)

        st.subheader("Download Data Dictionary")
        download_label = "Download Updated Data Dictionary" if rows_removed else "Download Data Dictionary"
        excel_file = io.BytesIO()
        st.session_state.balance_sheet_data.to_excel(excel_file, index=False)
        excel_file.seek(0)
        st.download_button(download_label, excel_file, "balance_sheet_data_dictionary.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") 

def cash_flow_statement_CF():
    # Initial setup
    cash_flow_data_dictionary_file = "cash_flow_data_dictionary.xlsx"  

    def load_lookup_table(filename):
        try:
            return pd.read_excel(filename)
        except FileNotFoundError:
            return pd.DataFrame(columns=['Account', 'Mnemonic', 'CIQ', 'Label'])

    if 'cash_flow_lookup_df' not in st.session_state:
        st.session_state.cash_flow_lookup_df = load_lookup_table(cash_flow_data_dictionary_file)

    def save_lookup_table(df, filename):
        df.to_excel(filename, index=False)

    def sort_by_label_and_account(df):
        logging.info("Sorting by label and account...")
        sort_order = {
            "Operating Activities": 0,
            "Investing Activities": 1,
            "Financing Activities": 2,
            "Cash Flow from Other": 3,
            "Supplemental Cash Flow": 4
        }

        df['Label_Order'] = df['Label'].map(sort_order)
        df['Total_Order'] = df['Account'].str.contains('Total', case=False).astype(int)

        df = df.sort_values(by=['Label_Order', 'Label', 'Total_Order', 'Account']).drop(columns=['Label_Order', 'Total_Order'])
        return df

    def sort_by_label_and_final_mnemonic(df):
        logging.info("Sorting by label and final mnemonic selection...")
        sort_order = {
            "Operating Activities": 0,
            "Investing Activities": 1,
            "Financing Activities": 2,
            "Cash Flow from Other": 3,
            "Supplemental Cash Flow": 4
        }

        df['Label_Order'] = df['Label'].map(sort_order)
        df['Total_Order'] = df['Final Mnemonic Selection'].str.contains('Total', case=False).astype(int)

        df = df.sort_values(by=['Label_Order', 'Total_Order', 'Final Mnemonic Selection']).drop(columns=['Label_Order', 'Total_Order'])
        return df

    st.title("CASH FLOW STATEMENT LTMA")

    tab1, tab2, tab3, tab4 = st.tabs(["Table Extractor", "Aggregate My Data", "Mappings and Data Consolidation", "Cash Flow Data Dictionary"])

    with tab1:
        uploaded_file = st.file_uploader("Choose a JSON file", type="json", key='json_uploader_cfs')
        if uploaded_file is not None:
            data = json.load(uploaded_file)
            st.warning("PLEASE NOTE: In the Setting Bounds Preview Window, you will see only your respective labels. In the Updated Columns Preview Window, you will see only your renamed column headers. The labels from the Setting Bounds section will not appear in the Updated Columns Preview.")
            st.warning("PLEASE ALSO NOTE: An Account column must also be designated when you are in the Rename Columns section.")

            tables = []
            for block in data['Blocks']:
                if block['BlockType'] == 'TABLE':
                    table = {}
                    if 'Relationships' in block:
                        for relationship in block['Relationships']:
                            if relationship['Type'] == 'CHILD':
                                for cell_id in relationship['Ids']:
                                    cell_block = next((b for b in data['Blocks'] if b['Id'] == cell_id), None)
                                    if cell_block:
                                        row_index = cell_block.get('RowIndex', 0)
                                        col_index = cell_block.get('ColumnIndex', 0)
                                        if row_index not in table:
                                            table[row_index] = {}
                                        cell_text = ''
                                        if 'Relationships' in cell_block:
                                            for rel in cell_block['Relationships']:
                                                if rel['Type'] == 'CHILD':
                                                    for word_id in rel['Ids']:
                                                        word_block = next((w for w in data['Blocks'] if w['Id'] == word_id), None)
                                                        if word_block and word_block['BlockType'] == 'WORD':
                                                            cell_text += ' ' + word_block.get('Text', '')
                                        table[row_index][col_index] = cell_text.strip()
                    table_df = pd.DataFrame.from_dict(table, orient='index').sort_index()
                    table_df = table_df.sort_index(axis=1)
                    tables.append(table_df)
            all_tables = pd.concat(tables, axis=0, ignore_index=True)
            if len(all_tables.columns) == 0:
                st.error("No columns found in the uploaded JSON file.")
                return

            column_a = all_tables.columns[0]
            all_tables.insert(0, 'Label', '')

            st.subheader("Data Preview")
            st.dataframe(all_tables)

            def get_unique_options(series):
                counts = series.value_counts()
                unique_options = []
                occurrence_counts = {}
                for idx, item in series.items():
                    if pd.isna(item) or item == '':
                        unique_options.append(f"[Blank {idx}]")
                    else:
                        if counts[item] > 1:
                            if item not in occurrence_counts:
                                occurrence_counts[item] = 1
                            else:
                                occurrence_counts[item] += 1
                            unique_options.append(f"{item} {occurrence_counts[item]}")
                        else:
                            unique_options.append(item)
                return unique_options

            labels = ["Operating Activities", "Investing Activities", "Financing Activities", "Cash Flow from Other", "Supplemental Cash Flow"]
            selections = []

            for label in labels:
                st.subheader(f"Setting bounds for {label}")
                options = [''] + get_unique_options(all_tables[column_a].fillna(''))
                start_label = st.selectbox(f"Start Label for {label}", options, key=f"start_{label}_cfs")
                end_label = st.selectbox(f"End Label for {label}", options, key=f"end_{label}_cfs")
                selections.append((label, start_label, end_label))

            new_column_names = {col: col for col in all_tables.columns}

            def update_labels(df):
                df['Label'] = ''
                account_column = new_column_names.get(column_a, column_a)
                
                # Insert index numbers in blank cells for easier identification
                df[account_column] = df[account_column].apply(lambda x: str(x) if pd.notna(x) else '')

                for index in df.index:
                    if df.at[index, account_column] == '':
                        df.at[index, account_column] = f"[Blank {index}]"
                
                for label, start_label, end_label in selections:
                    if start_label and end_label:
                        try:
                            # If start_label is a blank (indicated by index), use index directly
                            if start_label.startswith("[Blank"):
                                start_index = int(re.search(r"\d+", start_label).group())
                            else:
                                start_label_parts = start_label.split()
                                start_label_base = " ".join(start_label_parts[:-1]) if start_label_parts[-1].isdigit() else start_label
                                start_instance = int(start_label_parts[-1]) if start_label_parts[-1].isdigit() else 1
                                start_indices = df[df[account_column].str.contains(start_label_base, regex=False, na=False)].index
                                start_index = start_indices[start_instance - 1] if len(start_indices) >= start_instance else None
                            
                            # If end_label is a blank (indicated by index), use index directly
                            if end_label.startswith("[Blank"):
                                end_index = int(re.search(r"\d+", end_label).group())
                            else:
                                end_label_parts = end_label.split()
                                end_label_base = " ".join(end_label_parts[:-1]) if end_label_parts[-1].isdigit() else end_label
                                end_instance = int(end_label_parts[-1]) if end_label_parts[-1].isdigit() else 1
                                end_indices = df[df[account_column].str.contains(end_label_base, regex=False, na=False)].index
                                end_index = end_indices[end_instance - 1] if len(end_indices) >= end_instance else None
                            
                            # Ensure start and end indices are set, or skip if missing
                            if start_index is not None and end_index is not None:
                                df.loc[start_index:end_index, 'Label'] = label
                            else:
                                st.error(f"Invalid label bounds for {label}. Not enough instances found.")
                        except KeyError as e:
                            st.error(f"Error accessing column '{account_column}': {e}. Skipping...")
                    else:
                        st.info(f"No selections made for {label}. Skipping...")
                return df

            if st.button("Preview Setting Bounds ONLY", key="preview_setting_bounds_cfs"):
                preview_table = update_labels(all_tables.copy())
                st.subheader("Preview of Setting Bounds")
                st.dataframe(preview_table)

            st.subheader("Rename Columns")
            new_column_names = {}
            fiscal_year_options = [f"FY{year}" for year in range(2017, 2027)]
            ytd_options = [f"YTD{quarter}{year}" for year in range(2017, 2027) for quarter in range(1, 4)]
            dropdown_options = [''] + ['Account'] + fiscal_year_options + ytd_options

            for col in all_tables.columns:
                new_name_text = st.text_input(f"Rename '{col}' to:", value=col, key=f"rename_{col}_text_cfs")
                new_name_dropdown = st.selectbox(f"Or select predefined name for '{col}':", dropdown_options, key=f"rename_{col}_dropdown_cfs", index=0)
                new_column_names[col] = new_name_dropdown if new_name_dropdown else new_name_text
            
            all_tables.rename(columns=new_column_names, inplace=True)
            st.write("Updated Columns:", all_tables.columns.tolist())
            st.dataframe(all_tables)

            st.subheader("Edit Data Frame")
            editable_df = st.data_editor(all_tables)
            
            st.subheader("Select columns to keep before export")
            columns_to_keep = []
            for col in editable_df.columns:
                if st.checkbox(f"Keep column '{col}'", value=True, key=f"keep_{col}_cfs"):
                    columns_to_keep.append(col)

            st.subheader("Select numerical columns")
            numerical_columns = []
            for col in editable_df.columns:
                if st.checkbox(f"Numerical column '{col}'", value=False, key=f"num_{col}_cfs"):
                    numerical_columns.append(col)

            if 'Label' not in columns_to_keep:
                columns_to_keep.insert(0, 'Label')

            if 'Account' not in columns_to_keep:
                columns_to_keep.insert(1, 'Account')

            st.subheader("Convert Units")
            selected_columns = st.multiselect("Select columns for conversion", options=numerical_columns, key="columns_selection_cfs")
            selected_value = st.radio("Select conversion value", ["Actuals", "Thousands", "Millions", "Billions"], index=0, key="conversion_value_cfs")

            conversion_factors = {
                "Actuals": 1,
                "Thousands": 1000,
                "Millions": 1000000,
                "Billions": 1000000000
            }

            if st.button("Apply Selected Labels and Generate Excel", key="apply_selected_labels_generate_excel_tab1_cfs"):
                updated_table = update_labels(editable_df.copy())
                updated_table = updated_table[[col for col in columns_to_keep if col in updated_table.columns]]

                updated_table = updated_table[updated_table['Label'].str.strip() != '']
                updated_table = updated_table[updated_table['Account'].str.strip() != '']

                for col in numerical_columns:
                    if col in updated_table.columns:
                        updated_table[col] = updated_table[col].apply(clean_numeric_value)

                if selected_value != "Actuals":
                    updated_table = apply_unit_conversion(updated_table, selected_columns, conversion_factors[selected_value])

                updated_table.replace('-', 0, inplace=True)

                excel_file = io.BytesIO()
                updated_table.to_excel(excel_file, index=False)
                excel_file.seek(0)
                st.download_button("Download Excel", excel_file, "Table_Extractor_Cash_Flow_Statement.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab2:
        st.subheader("Aggregate My Data")
        
        uploaded_files = st.file_uploader("Upload your Excel files from Tab 1", type=['xlsx'], accept_multiple_files=True, key='xlsx_uploader_tab2_cfs')

        dfs = []
        if uploaded_files:
            dfs = [process_file(file) for file in uploaded_files if process_file(file) is not None]

        if dfs:
            combined_df = pd.concat(dfs, ignore_index=True)
            st.dataframe(combined_df)

            aggregated_table = aggregate_data(combined_df)
            aggregated_table = sort_by_label_and_account(aggregated_table)

            st.subheader("Aggregated Data")
            st.dataframe(aggregated_table)

            st.subheader("Preview Data and Edit Rows")
            zero_rows = check_all_zeroes(aggregated_table)
            zero_rows_indices = aggregated_table.index[zero_rows].tolist()
            st.write("Rows where all values (past the first 2 columns) are zero:", aggregated_table.loc[zero_rows_indices])
            
            edited_data = st.data_editor(aggregated_table, num_rows="dynamic")
            
            st.write("Highlighted rows with all zero values for potential removal:")
            for index in zero_rows_indices:
                st.write(f"Row {index}: {aggregated_table.loc[index].to_dict()}")
            
            rows_removed = False
            if st.button("Remove Highlighted Rows", key="remove_highlighted_rows_cfs"):
                aggregated_table = aggregated_table.drop(zero_rows_indices).reset_index(drop=True)
                rows_removed = True
                st.success("Highlighted rows removed successfully")
                st.dataframe(aggregated_table)

            st.subheader("Download Aggregated Data")
            download_label = "Download Updated Aggregated Excel" if rows_removed else "Download Aggregated Excel"
            excel_file = io.BytesIO()
            with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
                aggregated_table.to_excel(writer, sheet_name='Aggregated Data', index=False)
            excel_file.seek(0)
            st.download_button(download_label, excel_file, "Aggregate_My_Data_Cash_Flow_Statement.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("Please upload valid Excel files for aggregation.")

    with tab3:
        st.subheader("Mappings and Data Consolidation")

        uploaded_excel = st.file_uploader("Upload your Excel file for Mnemonic Mapping", type=['xlsx'], key='excel_uploader_tab3_cfs')

        currency_options = ["U.S. Dollar", "Euro", "British Pound Sterling", "Japanese Yen"]
        magnitude_options = ["Actuals", "Thousands", "Millions", "Billions", "Trillions"]

        selected_currency = st.selectbox("Select Currency", currency_options, key='currency_selection_tab3_cfs')
        selected_magnitude = st.selectbox("Select Magnitude", magnitude_options, key='magnitude_selection_tab3_cfs')
        company_name_cfs = st.text_input("Enter Company Name", key='company_name_input_cfs')

        if uploaded_excel is not None:
            df = pd.read_excel(uploaded_excel)

            statement_dates = {}
            for col in df.columns[2:]:
                statement_date = st.text_input(f"Enter statement date for {col}", key=f"statement_date_{col}_cfs")
                statement_dates[col] = statement_date

            st.write("Columns in the uploaded file:", df.columns.tolist())

            if 'Account' not in df.columns:
                st.error("The uploaded file does not contain an 'Account' column.")
            else:
                # Apply the new fuzzy matching function
                df['Mnemonic'] = ''
                df['Manual Selection'] = ''
                for idx, row in df.iterrows():
                    account_value = row['Account']
                    label_value = row.get('Label', '')
                    if pd.notna(account_value):
                        best_match, score = get_best_match(account_value, label_value, st.session_state.cash_flow_lookup_df)
                        if best_match is not None:
                            df.at[idx, 'Mnemonic'] = best_match['Mnemonic']
                            # Show match score for debugging
                            st.write(f"Match for '{account_value}': {best_match['Mnemonic']} (Score: {score:.2f})")
                        else:
                            df.at[idx, 'Mnemonic'] = 'Manual Mapping Required'
                            st.markdown(f"**Manual Mapping Required for:** {account_value} [{label_value} - Index {idx}]")

                for idx, row in df.iterrows():
                    account_value = row['Account']
                    label_value = row.get('Label', '')
                    
                    # If manual mapping required, show dropdown with all possible mnemonics
                    if row['Mnemonic'] == 'Manual Mapping Required':
                        # Get mnemonics for the specific label if available
                        if pd.notna(label_value) and label_value != '':
                            label_mnemonics = st.session_state.cash_flow_lookup_df[
                                st.session_state.cash_flow_lookup_df['Label'].str.strip().str.lower() == label_value.strip().lower()
                            ]['Mnemonic'].unique()
                            
                            # If no mnemonics for this label, show all mnemonics
                            if len(label_mnemonics) == 0:
                                label_mnemonics = st.session_state.cash_flow_lookup_df['Mnemonic'].unique()
                        else:
                            label_mnemonics = st.session_state.cash_flow_lookup_df['Mnemonic'].unique()
                        
                        manual_selection_options = list(label_mnemonics)
                        manual_selection = st.selectbox(
                            f"Select mapping for '{account_value}'",
                            options=[''] + manual_selection_options + ['REMOVE ROW', 'MANUAL OVERRIDE'],
                            key=f"select_{idx}_tab3_cfs"
                        )
                        if manual_selection:
                            df.at[idx, 'Manual Selection'] = manual_selection.strip()
                    else:
                        # If automatic mapping worked, still allow override
                        manual_override = st.checkbox(f"Override mapping for '{account_value}'", key=f"override_{idx}_tab3_cfs")
                        if manual_override:
                            label_mnemonics = st.session_state.cash_flow_lookup_df['Mnemonic'].unique()
                            manual_selection_options = list(label_mnemonics)
                            manual_selection = st.selectbox(
                                f"Select mapping for '{account_value}'",
                                options=[''] + manual_selection_options + ['REMOVE ROW'],
                                key=f"select_override_{idx}_tab3_cfs"
                            )
                            if manual_selection:
                                df.at[idx, 'Manual Selection'] = manual_selection.strip()

                st.dataframe(df[['Label', 'Account', 'Mnemonic', 'Manual Selection']])

                if st.button("Generate Excel with Lookup Results", key="generate_excel_lookup_results_tab3_cfs"):
                    df['Final Mnemonic Selection'] = df.apply(
                        lambda row: row['Manual Selection'].strip() if row['Manual Selection'].strip() != '' else row['Mnemonic'], 
                        axis=1
                    )
                    final_output_df = df[df['Manual Selection'] != 'REMOVE ROW'].copy()

                    combined_df = create_combined_df([final_output_df])
                    combined_df = sort_by_label_and_final_mnemonic(combined_df)

                    def lookup_ciq(mnemonic):
                        if mnemonic == 'Manual Mapping Required':
                            return 'CIQ ID Required'
                        ciq_value = st.session_state.cash_flow_lookup_df.loc[st.session_state.cash_flow_lookup_df['Mnemonic'] == mnemonic, 'CIQ']
                        if ciq_value.empty:
                            return 'CIQ ID Required'
                        return ciq_value.values[0]

                    combined_df['CIQ'] = combined_df['Final Mnemonic Selection'].apply(lookup_ciq)

                    columns_order = ['Label', 'Final Mnemonic Selection', 'CIQ'] + [col for col in combined_df.columns if col not in ['Label', 'Final Mnemonic Selection', 'CIQ']]
                    combined_df = combined_df[columns_order]

                    as_presented_df = final_output_df.drop(columns=['CIQ', 'Mnemonic', 'Manual Selection'], errors='ignore')
                    as_presented_columns_order = ['Label', 'Account', 'Final Mnemonic Selection'] + [col for col in as_presented_df.columns if col not in ['Label', 'Account', 'Final Mnemonic Selection']]
                    as_presented_df = as_presented_df[as_presented_columns_order]

                    excel_file = io.BytesIO()
                    with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
                        combined_df.to_excel(writer, sheet_name='Standardized - Cash Flow', index=False)
                        as_presented_df.to_excel(writer, sheet_name='As Presented - Cash Flow', index=False)
                        cover_df = pd.DataFrame({
                            'Selection': ['Currency', 'Magnitude', 'Company Name'] + list(statement_dates.keys()),
                            'Value': [selected_currency, selected_magnitude, company_name_cfs] + list(statement_dates.values())
                        })
                        cover_df.to_excel(writer, sheet_name='Cover', index=False)
                    excel_file.seek(0)
                    st.download_button("Download Excel", excel_file, "Mappings_and_Data_Consolidation_Cash_Flow_Statement.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                if st.button("Update Data Dictionary with Manual Mappings", key="update_data_dictionary_tab3_cfs"):
                    df['Final Mnemonic Selection'] = df.apply(
                        lambda row: row['Manual Selection'] if row['Manual Selection'] != '' else row['Mnemonic'], 
                        axis=1
                    )
                    new_entries = []
                    for idx, row in df.iterrows():
                        manual_selection = row['Manual Selection']
                        final_mnemonic = row['Final Mnemonic Selection']
                        ciq_value = st.session_state.cash_flow_lookup_df.loc[st.session_state.cash_flow_lookup_df['Mnemonic'] == final_mnemonic, 'CIQ'].values[0] if not st.session_state.cash_flow_lookup_df.loc[st.session_state.cash_flow_lookup_df['Mnemonic'] == final_mnemonic, 'CIQ'].empty else 'CIQ ID Required'

                        if manual_selection == 'REMOVE ROW':
                            pass  # Skip this row
                        elif manual_selection != '':
                            if row['Account'] not in st.session_state.cash_flow_lookup_df['Account'].values:
                                new_entries.append({'Account': row['Account'], 'Mnemonic': final_mnemonic, 'CIQ': ciq_value, 'Label': row['Label']})
                            else:
                                st.session_state.cash_flow_lookup_df.loc[st.session_state.cash_flow_lookup_df['Account'] == row['Account'], 'Mnemonic'] = final_mnemonic
                                st.session_state.cash_flow_lookup_df.loc[st.session_state.cash_flow_lookup_df['Account'] == row['Account'], 'Label'] = row['Label']
                                st.session_state.cash_flow_lookup_df.loc[st.session_state.cash_flow_lookup_df['Account'] == row['Account'], 'CIQ'] = ciq_value
                    if new_entries:
                        st.session_state.cash_flow_lookup_df = pd.concat([st.session_state.cash_flow_lookup_df, pd.DataFrame(new_entries)], ignore_index=True)
                    st.session_state.cash_flow_lookup_df.reset_index(drop=True, inplace=True)
                    save_lookup_table(st.session_state.cash_flow_lookup_df, cash_flow_data_dictionary_file)
                    st.success("Data Dictionary Updated Successfully")

    with tab4:
        st.subheader("Cash Flow Data Dictionary")

        uploaded_dict_file = st.file_uploader("Upload a new Data Dictionary Excel file", type=['xlsx'], key='dict_uploader_tab4_cfs')
        if uploaded_dict_file is not None:
            new_lookup_df = pd.read_excel(uploaded_dict_file)
            st.session_state.cash_flow_lookup_df = new_lookup_df
            save_lookup_table(st.session_state.cash_flow_lookup_df, cash_flow_data_dictionary_file)
            st.success("Data Dictionary uploaded and updated successfully!")

        st.dataframe(st.session_state.cash_flow_lookup_df)

        remove_indices = st.multiselect("Select rows to remove", st.session_state.cash_flow_lookup_df.index, key='remove_indices_tab4_cfs')
        rows_removed = False
        if st.button("Remove Selected Rows", key="remove_selected_rows_tab4_cfs"):
            st.session_state.cash_flow_lookup_df = st.session_state.cash_flow_lookup_df.drop(remove_indices).reset_index(drop=True)
            save_lookup_table(st.session_state.cash_flow_lookup_df, cash_flow_data_dictionary_file)
            rows_removed = True
            st.success("Selected rows removed successfully!")
            st.dataframe(st.session_state.cash_flow_lookup_df)

        st.subheader("Download Data Dictionary")
        download_label = "Download Updated Data Dictionary" if rows_removed else "Download Data Dictionary"
        excel_file = io.BytesIO()
        st.session_state.cash_flow_lookup_df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        st.download_button(download_label, excel_file, "cash_flow_data_dictionary.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
def income_statement():
    # Initial setup and data loading
    income_statement_data_dictionary_file = 'income_statement_data_dictionary.xlsx'

    def load_lookup_table(filename):
        try:
            return pd.read_excel(filename)
        except FileNotFoundError:
            return pd.DataFrame(columns=['Account', 'Mnemonic', 'CIQ'])

    if 'income_statement_lookup_df' not in st.session_state:
        st.session_state.income_statement_lookup_df = load_lookup_table(income_statement_data_dictionary_file)

    def sort_by_sort_index(df):
        if 'Sort Index' in df.columns:
            df = df.sort_values(by=['Sort Index'])
        return df

    def aggregate_data_IS(uploaded_files):
        dataframes = []
        unique_accounts = set()
        for file in uploaded_files:
            df = pd.read_excel(file)
            df.columns = [str(col).strip() for col in df.columns]
            if 'Account' not in df.columns:
                st.error(f"Column 'Account' not found in file {file.name}")
                return None
            df['Sort Index'] = range(1, len(df) + 1)
            dataframes.append(df)
            unique_accounts.update(df['Account'].dropna().unique())
        concatenated_df = pd.concat(dataframes, ignore_index=True)
        statement_date_rows = concatenated_df[concatenated_df['Account'].str.contains('Statement Date:', na=False)]
        numeric_rows = concatenated_df[~concatenated_df['Account'].str.contains('Statement Date:', na=False)]
        for col in numeric_rows.columns:
            if col not in ['Account', 'Sort Index', 'Positive Decreases NI']:
                numeric_rows[col] = numeric_rows[col].apply(clean_numeric_value)
        numeric_rows.fillna(0, inplace=True)
        for col in numeric_rows.columns:
            if col not in ['Account', 'Sort Index', 'Positive Decreases NI']:
                numeric_rows[col] = pd.to_numeric(numeric_rows[col], errors='coerce').fillna(0)
        aggregated_df = numeric_rows.groupby(['Account'], as_index=False).sum(min_count=1)
        statement_date_rows['Sort Index'] = 100
        statement_date_rows = statement_date_rows.groupby('Account', as_index=False).first()
        final_df = pd.concat([aggregated_df, statement_date_rows], ignore_index=True)
        final_df.insert(1, 'Positive Decreases NI', False)
        sort_index_column = final_df.pop('Sort Index')
        final_df['Sort Index'] = sort_index_column
        final_df.sort_values('Sort Index', inplace=True)
        return final_df

    def update_negative_values(df):
        criteria = [
            "IQ_COGS",
            "IQ_SGA_SUPPL",
            "IQ_RD_EXP",
            "IQ_DA_SUPPL",
            "IQ_STOCK_BASED",
            "IQ_OTHER_OPER",
            "IQ_INC_TAX"
        ]
        for index, row in df.iterrows():
            if row['CIQ'] in criteria:
                for col in df.columns[2:]:
                    if isinstance(row[col], (int, float)) and row[col] < 0:
                        df.at[index, col] = row[col] * 1
        return df

    # Conversion factors (was missing)
    conversion_factors = {
        "Actuals": 1,
        "Thousands": 1_000,
        "Millions": 1_000_000,
        "Billions": 1_000_000_000
    }

    st.title("INCOME STATEMENT LTMA")
    tab1, tab2, tab3, tab4 = st.tabs(["Table Extractor", "Aggregate My Data", "Mappings and Data Consolidation", "Income Statement Data Dictionary"])

    with tab1:
        uploaded_file = st.file_uploader("Choose a JSON file", type="json", key='json_uploader_is')
        if uploaded_file is not None:
            data = json.load(uploaded_file)
            tables = []
            for block in data['Blocks']:
                if block['BlockType'] == 'TABLE':
                    table = {}
                    if 'Relationships' in block:
                        for relationship in block['Relationships']:
                            if relationship['Type'] == 'CHILD':
                                for cell_id in relationship['Ids']:
                                    cell_block = next((b for b in data['Blocks'] if b['Id'] == cell_id), None)
                                    if cell_block:
                                        row_index = cell_block.get('RowIndex', 0)
                                        col_index = cell_block.get('ColumnIndex', 0)
                                        if row_index not in table:
                                            table[row_index] = {}
                                        cell_text = ''
                                        if 'Relationships' in cell_block:
                                            for rel in cell_block['Relationships']:
                                                if rel['Type'] == 'CHILD':
                                                    for word_id in rel['Ids']:
                                                        word_block = next((w for w in data['Blocks'] if w['Id'] == word_id), None)
                                                        if word_block and word_block['BlockType'] == 'WORD':
                                                            cell_text += ' ' + word_block.get('Text', '')
                                        table[row_index][col_index] = cell_text.strip()
                    table_df = pd.DataFrame.from_dict(table, orient='index').sort_index()
                    table_df = table_df.sort_index(axis=1)
                    tables.append(table_df)
            all_tables = pd.concat(tables, axis=0, ignore_index=True)
            column_a = all_tables.columns[0]

            st.subheader("Data Preview")
            st.dataframe(all_tables)

            st.subheader("Rename Columns")
            new_column_names = {}
            fiscal_year_options = [f"FY{year}" for year in range(2017, 2027)]
            ytd_options = [f"YTD{quarter}{year}" for year in range(2017, 2027) for quarter in range(1, 4)]
            dropdown_options = [''] + ['Account'] + fiscal_year_options + ytd_options

            for col in all_tables.columns:
                new_name_text = st.text_input(f"Rename '{col}' to:", value=col, key=f"rename_{col}_text_is")
                new_name_dropdown = st.selectbox(f"Or select predefined name for '{col}':", dropdown_options, key=f"rename_{col}_dropdown_is")
                new_column_names[col] = new_name_dropdown if new_name_dropdown else new_name_text

            all_tables.rename(columns=new_column_names, inplace=True)
            st.write("Updated Columns:", all_tables.columns.tolist())
            st.dataframe(all_tables)

            st.subheader("Edit and Remove Rows")
            editable_df = st.data_editor(all_tables, num_rows="dynamic", use_container_width=True)

            st.subheader("Select columns to keep before export")
            columns_to_keep = []
            for col in editable_df.columns:
                if st.checkbox(f"Keep column '{col}'", value=True, key=f"keep_{col}_tab1_is"):
                    columns_to_keep.append(col)

            editable_df = editable_df[columns_to_keep]

            st.subheader("Select numerical columns")
            numerical_columns = []
            for col in editable_df.columns:
                if st.checkbox(f"Numerical column '{col}'", value=False, key=f"num_{col}_is"):
                    numerical_columns.append(col)

            st.subheader("Convert Units")
            selected_columns = st.multiselect("Select columns for conversion", options=numerical_columns, key="columns_selection_is")
            selected_conversion_factor = st.radio("Select conversion factor", options=list(conversion_factors.keys()), key="conversion_factor_is")

            if st.button("Apply Selected Labels and Generate Excel", key="apply_selected_labels_generate_excel_tab1_is"):
                updated_table = editable_df

                for col in numerical_columns:
                    updated_table[col] = updated_table[col].apply(clean_numeric_value)

                if selected_conversion_factor and selected_conversion_factor in conversion_factors:
                    conversion_factor = conversion_factors[selected_conversion_factor]
                    updated_table = apply_unit_conversion(updated_table, selected_columns, conversion_factor)

                updated_table.replace('-', 0, inplace=True)

                excel_file = io.BytesIO()
                updated_table.to_excel(excel_file, index=False)
                excel_file.seek(0)
                st.download_button("Download Excel", excel_file, "Table_Extractor_Income_Statement.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


    with tab2:
        st.subheader("Aggregate My Data")

        uploaded_files = st.file_uploader("Upload Excel files", type=['xlsx'], accept_multiple_files=True, key='excel_uploader_amd_is')
        if uploaded_files:
            aggregated_df = aggregate_data_IS(uploaded_files)
            if aggregated_df is not None:
                st.subheader("Aggregated Data Preview")

                st.write("Select and order columns")
                columns_to_order = st.multiselect("Select columns in order", aggregated_df.columns.tolist(), default=aggregated_df.columns.tolist())
                
                if columns_to_order:
                    aggregated_df = aggregated_df[columns_to_order]

                editable_df = st.data_editor(aggregated_df, use_container_width=True)
                editable_df_excluded = editable_df.iloc[:-1]

                for col in editable_df_excluded.columns:
                    if col not in ['Account', 'Sort Index', 'Positive Decreases NI']:
                        editable_df_excluded[col] = pd.to_numeric(editable_df_excluded[col], errors='coerce').fillna(0)

                numeric_cols = editable_df_excluded.select_dtypes(include='number').columns.tolist()
                for index, row in editable_df_excluded.iterrows():
                    if row['Positive Decreases NI'] and row['Sort Index'] != 100:
                        for col in numeric_cols:
                            if col not in ['Sort Index']:
                                editable_df_excluded.at[index, col] = row[col] * -1

                final_df = pd.concat([editable_df_excluded, editable_df.iloc[-1:]], ignore_index=True)

                if 'Positive Decreases NI' in final_df.columns:
                    final_df.drop(columns=['Positive Decreases NI'], inplace=True)

                excel_file = io.BytesIO()
                final_df.to_excel(excel_file, index=False)
                excel_file.seek(0)
                st.download_button("Download Excel", excel_file, "Aggregate_My_Data_Income_Statement.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab3:
        st.subheader("Mappings and Data Consolidation")

        uploaded_excel_is = st.file_uploader("Upload your Excel file for Mnemonic Mapping", type=['xlsx'], key='excel_uploader_tab3_is')

        currency_options_is = ["U.S. Dollar", "Euro", "British Pound Sterling", "Japanese Yen"]
        magnitude_options_is = ["Actuals", "Thousands", "Millions", "Billions", "Trillions"]

        selected_currency_is = st.selectbox("Select Currency", currency_options_is, key='currency_selection_tab3_is')
        selected_magnitude_is = st.selectbox("Select Magnitude", magnitude_options_is, key='magnitude_selection_tab3_is')
        company_name_is = st.text_input("Enter Company Name", key='company_name_input_is')

        statement_dates = {}
        if uploaded_excel_is is not None:
            df_is = pd.read_excel(uploaded_excel_is)

            for col in df_is.columns:
                if col not in ['Account', 'Mnemonic', 'Manual Selection', 'Sort Index']:
                    statement_dates[col] = st.text_input(f"Enter statement date for {col}", key=f"statement_date_{col}_is")

            st.write("Columns in the uploaded file:", df_is.columns.tolist())

            if 'Account' not in df_is.columns:
                st.error("The uploaded file does not contain an 'Account' column.")
            else:
                if 'Sort Index' not in df_is.columns:
                    df_is['Sort Index'] = range(1, len(df_is) + 1)

                # Apply the new fuzzy matching function (without label since income statement typically doesn't use labels)
                df_is['Mnemonic'] = ''
                df_is['Manual Selection'] = ''
                for idx, row in df_is.iterrows():
                    account_value = row['Account']
                    if pd.notna(account_value):
                        best_match, score = get_best_match(account_value, '', st.session_state.income_statement_lookup_df, threshold=0.5)
                        if best_match is not None:
                            df_is.at[idx, 'Mnemonic'] = best_match['Mnemonic']
                            # Show match score for debugging
                            st.write(f"Match for '{account_value}': {best_match['Mnemonic']} (Score: {score:.2f})")
                        else:
                            df_is.at[idx, 'Mnemonic'] = 'Manual Mapping Required'
                            st.markdown(f"**Manual Mapping Required for:** {account_value} - Index {idx}")

                for idx, row in df_is.iterrows():
                    account_value = row['Account']
                    
                    # If manual mapping required, show dropdown with all possible mnemonics
                    if row['Mnemonic'] == 'Manual Mapping Required':
                        manual_selection_options = st.session_state.income_statement_lookup_df['Mnemonic'].unique()
                        manual_selection_is = st.selectbox(
                            f"Select category for '{account_value}'",
                            options=[''] + list(manual_selection_options) + ['REMOVE ROW', 'MANUAL OVERRIDE'],
                            key=f"select_{idx}_tab3_is"
                        )
                        if manual_selection_is:
                            df_is.at[idx, 'Manual Selection'] = manual_selection_is.strip()
                    else:
                        # If automatic mapping worked, still allow override
                        manual_override = st.checkbox(f"Override mapping for '{account_value}'", key=f"override_{idx}_tab3_is")
                        if manual_override:
                            manual_selection_options = st.session_state.income_statement_lookup_df['Mnemonic'].unique()
                            manual_selection_is = st.selectbox(
                                f"Select category for '{account_value}'",
                                options=[''] + list(manual_selection_options) + ['REMOVE ROW'],
                                key=f"select_override_{idx}_tab3_is"
                            )
                            if manual_selection_is:
                                df_is.at[idx, 'Manual Selection'] = manual_selection_is.strip()

                st.dataframe(df_is[['Account', 'Mnemonic', 'Manual Selection']])

                if st.button("Generate Excel with Lookup Results", key="generate_excel_lookup_results_tab3_is"):
                    df_is['Final Mnemonic Selection'] = df_is.apply(
                        lambda row: row['Manual Selection'] if row['Manual Selection'] not in ['REMOVE ROW', ''] else row['Mnemonic'], 
                        axis=1
                    )
                    final_output_df_is = df_is[df_is['Final Mnemonic Selection'].str.strip() != 'REMOVE ROW'].copy()

                    def lookup_ciq_is(mnemonic):
                        if mnemonic == 'Manual Mapping Required':
                            return 'CIQ ID Required'
                        ciq_value_is = st.session_state.income_statement_lookup_df.loc[st.session_state.income_statement_lookup_df['Mnemonic'] == mnemonic, 'CIQ']
                        if ciq_value_is.empty:
                            return 'CIQ ID Required'
                        return ciq_value_is.values[0]

                    final_output_df_is['CIQ'] = final_output_df_is['Final Mnemonic Selection'].apply(lookup_ciq_is)

                    # Identify numerical columns (excluding 'Sort Index')
                    numerical_columns = [col for col in final_output_df_is.columns if col not in ['Final Mnemonic Selection', 'CIQ', 'Sort Index', 'Account', 'Mnemonic', 'Manual Selection']]

                    # Select necessary columns and aggregate
                    columns_to_keep = ['Final Mnemonic Selection', 'CIQ'] + numerical_columns
                    aggregated_df = final_output_df_is[columns_to_keep].groupby(['Final Mnemonic Selection', 'CIQ']).sum().reset_index()

                    # Update negative values
                    aggregated_df = update_negative_values(aggregated_df)

                    # Prepare 'As Presented' sheet
                    as_presented_df_is = final_output_df_is.drop(columns=['Mnemonic', 'Manual Selection', 'Sort Index'], errors='ignore')
                    as_presented_columns_order_is = ['Account', 'Final Mnemonic Selection'] + [col for col in as_presented_df_is.columns if col not in ['Account', 'Final Mnemonic Selection', 'CIQ']]
                    as_presented_df_is = as_presented_df_is[as_presented_columns_order_is]

                    excel_file_is = io.BytesIO()
                    with pd.ExcelWriter(excel_file_is, engine='xlsxwriter') as writer:
                        aggregated_df.to_excel(writer, sheet_name='Standardized - Income Stmt', index=False)
                        as_presented_df_is.to_excel(writer, sheet_name='As Presented - Income Stmt', index=False)
                        cover_df_is = pd.DataFrame({
                            'Selection': ['Currency', 'Magnitude', 'Company Name'] + list(statement_dates.keys()),
                            'Value': [selected_currency_is, selected_magnitude_is, company_name_is] + list(statement_dates.values())
                        })
                        cover_df_is.to_excel(writer, sheet_name='Cover', index=False)
                    excel_file_is.seek(0)
                    st.download_button("Download Excel", excel_file_is, "Mappings_and_Data_Consolidation_Income_Statement.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                
                if st.button("Update Data Dictionary with Manual Mappings", key="update_data_dictionary_tab3_is"):
                    df_is['Final Mnemonic Selection'] = df_is.apply(
                        lambda row: row['Manual Selection'] if row['Manual Selection'] not in ['REMOVE ROW', ''] else row['Mnemonic'], 
                        axis=1
                    )
                    new_entries_is = []
                    for idx, row in df_is.iterrows():
                        manual_selection_is = row['Manual Selection']
                        final_mnemonic_is = row['Final Mnemonic Selection']
                        if manual_selection_is == 'REMOVE ROW':
                            continue
                        ciq_value_is = st.session_state.income_statement_lookup_df.loc[st.session_state.income_statement_lookup_df['Mnemonic'] == final_mnemonic_is, 'CIQ'].values[0] if not st.session_state.income_statement_lookup_df.loc[st.session_state.income_statement_lookup_df['Mnemonic'] == final_mnemonic_is, 'CIQ'].empty else 'CIQ ID Required'

                        if manual_selection_is not in ['REMOVE ROW', '']:
                            if row['Account'] not in st.session_state.income_statement_lookup_df['Account'].values:
                                new_entries_is.append({'Account': row['Account'], 'Mnemonic': final_mnemonic_is, 'CIQ': ciq_value_is})
                            else:
                                st.session_state.income_statement_lookup_df.loc[st.session_state.income_statement_lookup_df['Account'] == row['Account'], 'Mnemonic'] = final_mnemonic_is
                                st.session_state.income_statement_lookup_df.loc[st.session_state.income_statement_lookup_df['Account'] == row['Account'], 'CIQ'] = ciq_value_is
                    if new_entries_is:
                        st.session_state.income_statement_lookup_df = pd.concat([st.session_state.income_statement_lookup_df, pd.DataFrame(new_entries_is)], ignore_index=True)
                    st.session_state.income_statement_lookup_df.reset_index(drop=True, inplace=True)
                    save_lookup_table(st.session_state.income_statement_lookup_df, income_statement_data_dictionary_file)
                    st.success("Data Dictionary Updated Successfully")

    with tab4:
        st.subheader("Income Statement Data Dictionary")

        uploaded_dict_file_is = st.file_uploader("Upload a new Data Dictionary Excel file", type=['xlsx'], key='dict_uploader_tab4_is')
        if uploaded_dict_file_is is not None:
            new_lookup_df_is = pd.read_excel(uploaded_dict_file_is)
            st.session_state.income_statement_lookup_df = new_lookup_df_is
            save_lookup_table(new_lookup_df_is, income_statement_data_dictionary_file)
            st.success("Data Dictionary uploaded and updated successfully!")

        st.dataframe(st.session_state.income_statement_lookup_df)

        remove_indices_is = st.multiselect("Select rows to remove", st.session_state.income_statement_lookup_df.index, key='remove_indices_tab4_is')
        if st.button("Remove Selected Rows", key="remove_selected_rows_tab4_is"):
            st.session_state.income_statement_lookup_df = st.session_state.income_statement_lookup_df.drop(remove_indices_is).reset_index(drop=True)
            save_lookup_table(st.session_state.income_statement_lookup_df, income_statement_data_dictionary_file)
            st.success("Selected rows removed successfully!")
            st.dataframe(st.session_state.income_statement_lookup_df)

        if st.button("Download Data Dictionary", key="download_data_dictionary_tab4_is"):
            excel_file_is = io.BytesIO()
            st.session_state.income_statement_lookup_df.to_excel(excel_file_is, index=False)
            excel_file_is.seek(0)
            st.download_button("Download Excel", excel_file_is, "income_statement_data_dictionary.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def populate_ciq_template_pt():
    st.title("Populate CIQ Template")

    tab1, tab2 = st.tabs(["Annual Upload Template", "Quarterly Upload Template"])

    def process_template(template_type):
        unique_id = template_type.lower()

        uploaded_template = st.file_uploader(f"Upload CIQ Template ({template_type})", type=['xlsx', 'xlsm'], key=f"template_{unique_id}")
        uploaded_balance_sheet = st.file_uploader(f"Upload Completed Balance Sheet Data ({template_type})", type=['xlsx', 'xlsm'], key=f"balance_sheet_{unique_id}")
        uploaded_cash_flow = st.file_uploader(f"Upload Completed Cash Flow Statement ({template_type})", type=['xlsx', 'xlsm'], key=f"cash_flow_{unique_id}")
        uploaded_income_statement = st.file_uploader(f"Upload Completed Income Statement Data ({template_type})", type=['xlsx', 'xlsm'], key=f"income_statement_{unique_id}")

        if st.button(f"Populate {template_type} Template Now", key=f"populate_button_{unique_id}") and uploaded_template and (uploaded_balance_sheet or uploaded_cash_flow or uploaded_income_statement):
            try:
                # Read the uploaded template file
                template_file = uploaded_template.read()
                try:
                    template_wb = load_workbook(BytesIO(template_file), keep_vba=True)
                except Exception as e:
                    st.error(f"Error loading template file: {e}")
                    return

                def process_sheet(sheet_file, sheet_name, row_range, date_row):
                    try:
                        sheet_wb = load_workbook(BytesIO(sheet_file), keep_vba=True)
                    except Exception as e:
                        st.error(f"Error loading {sheet_name} file: {e}")
                        return
                    
                    as_presented_sheet = sheet_wb[f"As Presented - {sheet_name}"]
                    standardized_sheet = pd.read_excel(BytesIO(sheet_file), sheet_name=f"Standardized - {sheet_name}")

                    # Check for required columns in the Standardized sheet
                    if 'CIQ' not in standardized_sheet.columns:
                        st.error(f"The column 'CIQ' is missing from the Standardized - {sheet_name}.")
                        return

                    st.write(f"{sheet_name} CIQ column found, proceeding...")
                    st.write(standardized_sheet.head())

                    # Copy the "As Presented" sheet to the template workbook
                    if f"As Presented - {sheet_name}" in template_wb.sheetnames:
                        del template_wb[f"As Presented - {sheet_name}"]
                    new_sheet = template_wb.create_sheet(f"As Presented - {sheet_name}")
                    for row in as_presented_sheet.iter_rows():
                        for cell in row:
                            new_sheet[cell.coordinate].value = cell.value

                    # Color the "As Presented" tab orange
                    tab_color = "FFA500"
                    new_sheet.sheet_properties.tabColor = tab_color

                    # Copy the "Standardized" sheet to the template workbook
                    if f"Standardized - {sheet_name}" in template_wb.sheetnames:
                        del template_wb[f"Standardized - {sheet_name}"]
                    standardized_ws = template_wb.create_sheet(f"Standardized - {sheet_name}")

                    # Write the header row
                    for col_num, header in enumerate(standardized_sheet.columns, 1):
                        standardized_ws.cell(row=1, column=col_num, value=header)

                    # Write the data rows
                    for r_idx, row in enumerate(standardized_sheet.itertuples(index=False), 2):
                        for c_idx, value in enumerate(row, 1):
                            standardized_ws.cell(row=r_idx, column=c_idx, value=value)

                    # Perform lookups and update the "Upload" sheet
                    upload_sheet = template_wb["Upload"]
                    ciq_values = standardized_sheet['CIQ'].tolist()
                    dates = list(standardized_sheet.columns[2:])  # Skip 'CIQ' and 'Final Mnemonic Selection'

                    st.write(f"CIQ Values from {sheet_name}:", ciq_values)
                    st.write(f"Dates from {sheet_name}:", dates)

                    if template_type == "Annual":
                        for row in upload_sheet.iter_rows(min_row=row_range[0], max_row=row_range[1], min_col=4, max_col=upload_sheet.max_column):
                            ciq_cell = upload_sheet.cell(row=row[0].row, column=11)
                            ciq_value = ciq_cell.value
                            if ciq_value in ciq_values:
                                st.write(f"Processing CIQ Value: {ciq_value} at row {row[0].row}")
                                for col in range(4, 10):
                                    date_value = upload_sheet.cell(row=date_row, column=col).value
                                    st.write(f"Checking date {date_value} at column {col}")
                                    if date_value in dates:
                                        lookup_value = standardized_sheet.loc[standardized_sheet['CIQ'] == ciq_value, date_value].sum()
                                        st.write(f"Lookup value for CIQ {ciq_value} and date {date_value}: {lookup_value}")
                                        if not pd.isna(lookup_value):
                                            cell_to_update = upload_sheet.cell(row=row[0].row, column=col)
                                            if cell_to_update.data_type == 'f' or cell_to_update.value is None:
                                                cell_to_update.value = lookup_value
                                                st.write(f"Updated {cell_to_update.coordinate} with value {lookup_value}")

                        for row in upload_sheet.iter_rows(min_row=row_range[1] + 1, max_row=row_range[1] + 1, min_col=4, max_col=9):
                            for cell in row:
                                if cell.value is not None:
                                    try:
                                        cell_value = float(cell.value)
                                        cell.value = -abs(cell_value)
                                    except ValueError:
                                        st.warning(f"Non-numeric value found in cell {cell.coordinate}, skipping negation.")
                    
                    elif template_type == "Quarterly":
                        for row in upload_sheet.iter_rows(min_row=row_range[0], max_row=row_range[1], min_col=4, max_col=21):
                            ciq_cell = upload_sheet.cell(row=row[0].row, column=23)
                            ciq_value = ciq_cell.value
                            if ciq_value in ciq_values:
                                st.write(f"Processing CIQ Value: {ciq_value} at row {row[0].row}")
                                for col in range(4, 22):
                                    date_value = upload_sheet.cell(row=date_row, column=col).value
                                    st.write(f"Checking date {date_value} at column {col}")
                                    if date_value in dates:
                                        lookup_value = standardized_sheet.loc[standardized_sheet['CIQ'] == ciq_value, date_value].sum()
                                        st.write(f"Lookup value for CIQ {ciq_value} and date {date_value}: {lookup_value}")
                                        if not pd.isna(lookup_value):
                                            cell_to_update = upload_sheet.cell(row=row[0].row, column=col)
                                            if cell_to_update.data_type == 'f' or cell_to_update.value is None:
                                                cell_to_update.value = lookup_value
                                                st.write(f"Updated {cell_to_update.coordinate} with value {lookup_value}")

                        for row in upload_sheet.iter_rows(min_row=row_range[1] + 1, max_row=row_range[1] + 1, min_col=4, max_col=21):
                            for cell in row:
                                if cell.value is not None:
                                    try:
                                        cell_value = float(cell.value)
                                        cell.value = -abs(cell_value)
                                    except ValueError:
                                        st.warning(f"Non-numeric value found in cell {cell.coordinate}, skipping negation.")

                # Process sheets based on the uploaded files
                if template_type == "Annual":
                    if uploaded_balance_sheet:
                        balance_sheet_file = uploaded_balance_sheet.read()
                        process_sheet(balance_sheet_file, "Balance Sheet", (96, 162), 94)
                    if uploaded_cash_flow:
                        cash_flow_file = uploaded_cash_flow.read()
                        process_sheet(cash_flow_file, "Cash Flow", (171, 234), 169)
                    if uploaded_income_statement:
                        income_statement_file = uploaded_income_statement.read()
                        process_sheet(income_statement_file, "Income Stmt", (14, 72), 12)

                if template_type == "Quarterly":
                    if uploaded_balance_sheet:
                        balance_sheet_file = uploaded_balance_sheet.read()
                        process_sheet(balance_sheet_file, "Balance Sheet", (96, 162), 94)
                    if uploaded_cash_flow:
                        cash_flow_file = uploaded_cash_flow.read()
                        process_sheet(cash_flow_file, "Cash Flow", (171, 234), 169)
                    if uploaded_income_statement:
                        income_statement_file = uploaded_income_statement.read()
                        process_sheet(income_statement_file, "Income Stmt", (14, 72), 12)

                # Save the updated workbook to a BytesIO object
                output = BytesIO()
                template_wb.save(output)
                template_data = output.getvalue()

                # Provide a download button for the updated template
                st.download_button(
                    label=f"Download Updated {template_type} Template",
                    data=template_data,
                    file_name=uploaded_template.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.success(f"{template_type} Template populated successfully. You can now download the updated template.")

            except Exception as e:
                st.error(f"An error occurred: {e}")

    with tab1:
        process_template("Annual")

    with tab2:
        process_template("Quarterly")
        
def backup_data_dictionaries():
    if st.button("Backup Data Dictionaries"):
        try:
            # Get the directory of the current script
            current_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Construct full file paths
            balance_sheet_path = os.path.join(current_dir, 'balance_sheet_data_dictionary.xlsx')
            cash_flow_path = os.path.join(current_dir, 'cash_flow_data_dictionary.xlsx')
            income_statement_path = os.path.join(current_dir, 'income_statement_data_dictionary.xlsx')
            
            # Load data dictionaries
            dfs = {}
            for name, path in [('Balance Sheet', balance_sheet_path), 
                               ('Cash Flow', cash_flow_path), 
                               ('Income Statement', income_statement_path)]:
                try:
                    if path.endswith('.xlsx'):
                        dfs[name] = pd.read_excel(path)
                    elif path.endswith('.csv'):
                        dfs[name] = pd.read_csv(path)
                    else:
                        st.warning(f"Unsupported file format for {name}. Skipping.")
                except FileNotFoundError:
                    st.warning(f"{name} dictionary file not found. Skipping.")
                except Exception as e:
                    st.warning(f"Error reading {name} dictionary: {str(e)}. Skipping.")
            
            if not dfs:
                st.error("No data dictionaries could be loaded. Please check the file paths and formats.")
                return
            
            # Create a new Excel writer object
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                for name, df in dfs.items():
                    df.to_excel(writer, sheet_name=name, index=False)
            
            output.seek(0)
            
            st.download_button(
                label="Download Backup",
                data=output,
                file_name="data_dictionaries_backup.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.success("Backup created successfully!")

        except Exception as e:
            st.error(f"An unexpected error occurred: {str(e)}")

def reset_aws_process():
    st.session_state.processed = False
    st.session_state.uploaded_file = None
    st.session_state.response_json_path = None
    st.session_state.simplified_response = None
    st.session_state.tables = None

def aws_textract_tab():
    st.title("AWS Textract with Streamlit - Table Extraction")
    st.write("Enter your AWS credentials and upload an image or PDF file to extract tables using AWS Textract.")

    if 'credentials_valid' not in st.session_state:
        st.session_state.credentials_valid = False

    if st.button("Confirm Credentials"):
        credentials_valid = check_aws_credentials()
        if credentials_valid:
            st.success("AWS credentials are valid!")
            st.session_state.credentials_valid = True
        else:
            st.error("Invalid AWS credentials. Please check and try again.")
            st.session_state.credentials_valid = False

    # File Upload (only show if credentials are valid)
    if st.session_state.credentials_valid:
        uploaded_file = st.file_uploader("Choose an image or PDF file", type=["jpg", "jpeg", "png", "pdf"])

        if uploaded_file is not None and (not st.session_state.get('processed', False) or uploaded_file != st.session_state.get('uploaded_file')):
            st.session_state.uploaded_file = uploaded_file
            temp_file_path = None
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[1]) as temp_file:
                    temp_file.write(uploaded_file.getvalue())
                    temp_file_path = temp_file.name
                
                textract_client = boto3.client('textract',
                                               aws_access_key_id=st.secrets["aws"]["AWS_ACCESS_KEY_ID"],
                                               aws_secret_access_key=st.secrets["aws"]["AWS_SECRET_ACCESS_KEY"],
                                               region_name=st.secrets["aws"]["AWS_REGION"])
                
                s3_client = boto3.client('s3',
                                         aws_access_key_id=st.secrets["aws"]["AWS_ACCESS_KEY_ID"],
                                         aws_secret_access_key=st.secrets["aws"]["AWS_SECRET_ACCESS_KEY"],
                                         region_name=st.secrets["aws"]["AWS_REGION"])
                
                with st.spinner("Processing document..."):
                    tables, response_json_path, simplified_response = process_document(temp_file_path, textract_client, s3_client, st.secrets["aws"]["S3_BUCKET_NAME"])
                
                st.session_state.processed = True
                st.session_state.response_json_path = response_json_path
                st.session_state.simplified_response = simplified_response
                st.session_state.tables = tables

            except botocore.exceptions.ClientError as e:
                st.error(f"AWS Error: {str(e)}")
            except Exception as e:
                st.error(f"An error occurred: {str(e)}")
                st.error("Please check the AWS credentials, S3 bucket permissions, and try again.")
            finally:
                # Clean up the temporary files
                if temp_file_path and os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)

    # Show download button and table extraction results only if processing is complete
    if st.session_state.get('processed', False):
        response_json_path = st.session_state.response_json_path
        simplified_response = st.session_state.simplified_response
        tables = st.session_state.tables
        
        # JSON file rename input
        st.subheader("Download Full JSON Response:")
        json_file_name = st.text_input("Enter JSON file name (without .json extension)", value='textract_response')
        
        if st.button("Download JSON"):
            with open(response_json_path, 'rb') as f:
                st.download_button(
                    label="Click to Download",
                    data=f,
                    file_name=f"{json_file_name}.json" if json_file_name else 'textract_response.json',
                    mime='application/json'
                )

        # Now display the extracted information
        st.subheader("Detected Tables:")
        st.write(f"Number of tables detected: {len(tables)}")
        if tables:
            for i, table in enumerate(tables):
                st.write(f"Table {i+1}:")
                if table:
                    df = pd.DataFrame(table)
                    st.dataframe(df)
                else:
                    st.write("Empty table detected")
        else:
            st.write("No tables detected")

        # Debug information
        st.subheader("Debug Information:")
        st.write(f"Number of blocks processed: {len(simplified_response['Blocks'])}")
        st.json(simplified_response)

        # Display structure of the first few blocks
        st.subheader("Structure of First Few Blocks:")
        first_few_blocks = simplified_response['Blocks'][:10]  # Display first 10 blocks
        for i, block in enumerate(first_few_blocks):
            st.write(f"Block {i}:")
            st.json(block)

        # Add a button to reset the process
        if st.button("Process another document"):
            reset_aws_process()
            st.rerun()

def extras_tab():
    st.title("Extras")
    tabs = st.tabs(["Backup Data Dictionaries", "AWS Textract"])

    with tabs[0]:
        backup_data_dictionaries()
    
    with tabs[1]:
        aws_textract_tab()
        
def check_aws_credentials():
    try:
        if "aws" in st.secrets:
            required_keys = ["AWS_ACCESS_KEY_ID", "AWS_SECRET_ACCESS_KEY", "AWS_REGION"]
            for key in required_keys:
                if key not in st.secrets["aws"]:
                    st.error(f"'{key}' not found in AWS secrets")
                    return False
            
            session = boto3.Session(
                aws_access_key_id=st.secrets["aws"]["AWS_ACCESS_KEY_ID"],
                aws_secret_access_key=st.secrets["aws"]["AWS_SECRET_ACCESS_KEY"],
                region_name=st.secrets["aws"]["AWS_REGION"]
            )
            sts = session.client('sts')
            sts.get_caller_identity()
            return True
        else:
            st.error("'aws' key not found in secrets")
            return False
    except Exception as e:
        st.error(f"Error checking AWS credentials: {str(e)}")
        return False

def safe_get(dict_obj, key, default=None):
    """Safely get a value from a dictionary."""
    return dict_obj.get(key, default)

def extract_table_data(table_blocks, blocks_map):
    rows = []
    for relationship in safe_get(table_blocks, 'Relationships', []):
        if safe_get(relationship, 'Type') == 'CHILD':
            for child_id in safe_get(relationship, 'Ids', []):
                cell = blocks_map.get(child_id, {})
                if safe_get(cell, 'BlockType') == 'CELL':
                    row_index = safe_get(cell, 'RowIndex', 1) - 1
                    col_index = safe_get(cell, 'ColumnIndex', 1) - 1
                    while len(rows) <= row_index:
                        rows.append([])
                    while len(rows[row_index]) <= col_index:
                        rows[row_index].append('')
                    cell_text = safe_get(cell, 'Text', '')
                    if not cell_text:
                        # If 'Text' is not present, try to get it from child relationships
                        for cell_relationship in safe_get(cell, 'Relationships', []):
                            if safe_get(cell_relationship, 'Type') == 'CHILD':
                                for word_id in safe_get(cell_relationship, 'Ids', []):
                                    word = blocks_map.get(word_id, {})
                                    cell_text += safe_get(word, 'Text', '') + ' '
                        cell_text = cell_text.strip()
                    rows[row_index][col_index] = cell_text
    return rows

def upload_to_s3(s3_client, file_bytes, bucket_name, object_name):
    try:
        s3_client.put_object(Body=file_bytes, Bucket=bucket_name, Key=object_name)
        st.success(f"Successfully uploaded file to S3: {bucket_name}/{object_name}")
    except botocore.exceptions.ClientError as e:
        st.error(f"Failed to upload file to S3: {str(e)}")
        raise

def start_document_analysis(textract_client, bucket_name, object_name):
    try:
        response = textract_client.start_document_analysis(
            DocumentLocation={'S3Object': {'Bucket': bucket_name, 'Name': object_name}},
            FeatureTypes=['TABLES']
        )
        return response['JobId']
    except botocore.exceptions.ClientError as e:
        st.error(f"Failed to start document analysis: {str(e)}")
        raise

def get_document_analysis(textract_client, job_id):
    pages = []
    next_token = None
    max_attempts = 60  # Increase this for larger documents
    attempt = 0
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    while attempt < max_attempts:
        try:
            if next_token:
                response = textract_client.get_document_analysis(JobId=job_id, NextToken=next_token)
            else:
                response = textract_client.get_document_analysis(JobId=job_id)
            
            status = response['JobStatus']
            status_text.text(f"Processing document... Status: {status}")
            
            if status == 'SUCCEEDED':
                pages.append(response)
                next_token = response.get('NextToken')
                if not next_token:
                    status_text.text("Document processing complete!")
                    progress_bar.progress(1.0)
                    break
            elif status == 'FAILED':
                raise Exception(f"Textract job failed: {response.get('StatusMessage')}")
            else:  # IN_PROGRESS
                time.sleep(5)
                attempt += 1
                progress_bar.progress(attempt / max_attempts)
        except Exception as e:
            st.error(f"An error occurred while getting document analysis: {str(e)}")
            raise
    
    if attempt >= max_attempts:
        raise Exception("Textract job timed out")
    
    return pages

def process_document(file_path, textract_client, s3_client, bucket_name):
    try:
        with open(file_path, 'rb') as document:
            file_bytes = document.read()
        
        _, file_extension = os.path.splitext(file_path)
        object_name = os.path.basename(file_path)
        
        if file_extension.lower() == '.pdf':
            st.info("Uploading document to S3...")
            upload_to_s3(s3_client, file_bytes, bucket_name, object_name)
            st.info("Starting Textract job...")
            job_id = start_document_analysis(textract_client, bucket_name, object_name)
            st.info(f"Textract job started with ID: {job_id}")
            response_pages = get_document_analysis(textract_client, job_id)
        else:
            response = textract_client.analyze_document(
                Document={'Bytes': file_bytes},
                FeatureTypes=['TABLES']
            )
            response_pages = [response]
        
        if not response_pages:
            raise Exception("Document analysis failed")
        
        # Extract and combine 'Blocks' from all pages
        all_blocks = []
        for page in response_pages:
            all_blocks.extend(page.get('Blocks', []))
        
        # Create a simplified JSON structure
        simplified_response = {'Blocks': all_blocks}
        
        # Save the simplified response as a JSON file
        response_json_path = file_path + '.json'
        with open(response_json_path, 'w') as json_file:
            json.dump(simplified_response, json_file, indent=4)
        
        # Extract tables
        tables = []

        # Process all blocks
        blocks_map = {safe_get(block, 'Id'): block for block in all_blocks}
        for block in all_blocks:
            block_type = safe_get(block, 'BlockType')
            if block_type == 'TABLE':
                tables.append(extract_table_data(block, blocks_map))
        
        return tables, response_json_path, simplified_response

    except Exception as e:
        st.error(f"An error occurred during document processing: {str(e)}")
        raise

def main():
    st.sidebar.title("Navigation")

    # HTML code to style the image position in the sidebar
    logo_html = """
        <div style="position: fixed; 
                    bottom: 30px; 
                    left: 30px;">
            <img src="https://raw.githubusercontent.com/akeenantexasfcs/ltma/main/TFCLogo.jpg" width="250"/>
        </div>
        """
    # Display the logo in the sidebar using HTML
    st.sidebar.markdown(logo_html, unsafe_allow_html=True)

    selection = st.sidebar.radio("Go to", ["Balance Sheet", "Cash Flow Statement", "Income Statement", "Populate CIQ Template", "Extras"])

    if selection == "Balance Sheet":
        balance_sheet_BS()
    elif selection == "Cash Flow Statement":
        cash_flow_statement_CF()
    elif selection == "Income Statement":
        income_statement()
    elif selection == "Populate CIQ Template":
        populate_ciq_template_pt()
    elif selection == "Extras":
        extras_tab()

if __name__ == '__main__':
    main()

